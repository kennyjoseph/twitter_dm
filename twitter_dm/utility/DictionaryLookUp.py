"""
This file is chiefly useful if you're using a thesaurus and you want to make submissions.

Get a hold of me if this is you and I can send some example code
"""
__author__ = 'kjoseph'

import codecs
from collections import defaultdict
from openpyxl import load_workbook

class DictionaryLookUp:
    def __init__(self,
                 file_location,
                 map_from_column=0,
                 map_to_column=1,
                 sep=",",
                 has_header=False,
                 all_lower=True):
        """
        :param file_location: location of the file containing the dictionary
        :param map_from_column: what to map from
        :param map_to_column: what to map to
        :param sep: separator splitting the from/to columns
        :param has_header: does the file have a header?
        :param all_lower: should everything be lowercased for the matching process?
        :return:
        """
        self.file_location = file_location
        self.mappings_dict = defaultdict(set)
        self.init_mappings(file_location,map_from_column,map_to_column,sep,has_header,all_lower)



    def insert_into_mappings(self,data, map_from_column,map_to_column,all_lower ):
        split_len = len(data)
        if split_len <= map_from_column or split_len <= map_to_column:
            print u'Line: {0} is malformatted'.format(",".join(data))
            raise Exception("malformed line")

        src = data[map_from_column]

        dest = ""
        if map_to_column != -1:
            dest = data[map_to_column]

        if src is None or dest is None:
            return False

        #dest = dest.replace("_","")

        src = src.strip()

        if all_lower:
            src = src.lower()

        self.mappings_dict[src].add(dest)
        return True

    '''
        Return mappings from surface form to concept.
        If map_to_column == -1, assumes there is no mapping, just replaces spaces w/ _
        If has_header, skips first line
        The sep param specifies the token used to separate the data
    '''
    def init_mappings(self,file_location,map_from_column,map_to_column,sep,has_header,all_lower):

        print 'getting mappings\n\tfile: {0}\n\theader: {1}\n\tsep: {2}\n\tmap_from: {3},\n\tmap_to: {4}'.format(
            file_location, has_header, sep,map_from_column,map_to_column
        )


        num_decode_failed = 0

        if '.xlsx' in file_location:
            input_file = load_workbook(filename = file_location, use_iterators = True)
            print 'loaded'
            input_sheet = input_file.worksheets[len(input_file.worksheets)-1]
            for row in input_sheet.iter_rows():
                val = [cell.value for cell in row]
                v = self.insert_into_mappings(val,map_from_column,map_to_column,all_lower)
                if not v:
                    num_decode_failed +=1

        else:
            input_file = codecs.open(file_location,"r","utf8",errors='replace')

            if has_header:
                input_file.readline()

            for line in input_file:
                v = self.insert_into_mappings(line.strip().split(sep),map_from_column,map_to_column,all_lower)
                if not v:
                    num_decode_failed +=1

        s = set()
        for k in self.mappings_dict.values():
            for v in k:
                s.add(v)
        print 'done! Size: {0}, N Unique Values: {1},  DECODE FAILED COUNT: {2}'.format(
            len(self.mappings_dict),len(s),num_decode_failed)


    def get_entities_for_text(self, text):
        if text in self.mappings_dict:
            return self.mappings_dict[text]
        return None






